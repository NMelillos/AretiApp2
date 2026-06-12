from io import BytesIO
import re
import textwrap

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


REPORT_GROUP_COLUMN = "report_group"
UNASSIGNED_GROUP = "Unassigned reporting group"


def _clean_text(value):
    if pd.isna(value):
        return ""
    return str(value).strip()


def _is_own_funds(value):
    return _clean_text(value).casefold() == "own funds"


def _looks_like_income_or_deposit(category, subcategory="", report_group=""):
    text = " ".join([
        _clean_text(category),
        _clean_text(subcategory),
        _clean_text(report_group),
    ]).casefold()
    if not text:
        return False
    income_markers = [
        "income",
        "deposit",
        "deposits",
        "revenue",
        "interest earned",
        "incoming",
    ]
    return any(marker in text for marker in income_markers)


def _ordered_unique(values):
    seen = set()
    ordered = []
    for value in values:
        text = str(value).strip()
        key = text.casefold()
        if not text or key in seen:
            continue
        seen.add(key)
        ordered.append(text)
    return ordered


def get_report_groups(categories):
    if categories.empty or REPORT_GROUP_COLUMN not in categories.columns:
        return []
    groups = (
        categories[REPORT_GROUP_COLUMN]
        .fillna("")
        .astype(str)
        .str.strip()
        .drop_duplicates()
        .tolist()
    )
    return [group for group in groups if group and not group.lower().startswith("0-")]


def _category_group_map(categories):
    if categories.empty or "category" not in categories.columns:
        return {}
    mapping = {}
    for _, row in categories.iterrows():
        category = str(row.get("category", "")).strip()
        if not category:
            continue
        group = str(row.get(REPORT_GROUP_COLUMN, "")).strip() if REPORT_GROUP_COLUMN in categories.columns else ""
        mapping.setdefault(category.casefold(), group)
    return mapping


def _month_context(expenses):
    end_month = expenses["month"].max() if not expenses.empty else pd.Timestamp.now().to_period("M")
    months = pd.period_range(end=end_month, periods=12, freq="M")
    labels = {month: month.to_timestamp().strftime("%Y-%m") for month in months}
    return months, labels


def _prepare_report_data(transactions, categories, report_group=None, include_own_funds=False, include_all_valid=False):
    tx = transactions.copy()
    for column in ["txn_date", "amount", "amount_usd", "currency", "category", "subcategory"]:
        if column not in tx.columns:
            tx[column] = ""

    if tx.empty:
        tx["report_amount"] = []
        return tx, tx.copy(), *(_month_context(tx))

    tx["txn_date"] = pd.to_datetime(tx["txn_date"], errors="coerce")
    tx["category"] = tx["category"].fillna("").astype(str).str.strip()
    tx["subcategory"] = tx["subcategory"].fillna("").astype(str).str.strip()
    tx["currency"] = tx["currency"].fillna("").astype(str).str.upper().str.strip()
    amount_numeric = pd.to_numeric(tx["amount"], errors="coerce")
    amount_usd_numeric = pd.to_numeric(tx["amount_usd"], errors="coerce")
    zero_usd_needs_attention = (
        amount_usd_numeric.fillna(0).abs().le(0.005)
        & amount_numeric.fillna(0).abs().gt(0.005)
    )
    tx["report_amount"] = amount_usd_numeric
    tx.loc[zero_usd_needs_attention & ~tx["currency"].eq("USD"), "report_amount"] = pd.NA
    usd_fallback = (
        (tx["report_amount"].isna() | zero_usd_needs_attention)
        & tx["currency"].eq("USD")
        & amount_numeric.notna()
    )
    tx.loc[usd_fallback, "report_amount"] = amount_numeric.loc[usd_fallback]
    tx["report_amount"] = tx["report_amount"].fillna(0)
    tx = tx.dropna(subset=["txn_date"]).copy()

    group_map = _category_group_map(categories)
    tx["report_group"] = tx["category"].map(lambda value: group_map.get(str(value).casefold(), ""))
    tx["report_group"] = tx["report_group"].replace("", UNASSIGNED_GROUP)

    positive_amount = tx["report_amount"] > 0
    income_like = tx.apply(
        lambda row: _looks_like_income_or_deposit(
            row.get("category", ""),
            row.get("subcategory", ""),
            row.get("report_group", ""),
        ),
        axis=1,
    )
    own_funds_mask = tx["category"].str.strip().str.casefold() == "own funds"
    expense_mask = (
        (tx["report_amount"] < 0)
        | (positive_amount & ~income_like & tx["report_group"].ne(UNASSIGNED_GROUP))
    )
    if include_all_valid:
        expenses = tx.copy()
    elif include_own_funds:
        expenses = tx[expense_mask | own_funds_mask].copy()
    else:
        expenses = tx[expense_mask & ~own_funds_mask].copy()
    expenses["expense_usd"] = expenses["report_amount"].abs()
    expenses["month"] = expenses["txn_date"].dt.to_period("M")

    if report_group:
        expenses = expenses[expenses["report_group"] == report_group].copy()

    months, labels = _month_context(expenses)
    return tx, expenses, months, labels


def _month_totals(frame, months):
    if frame.empty:
        return {month: 0.0 for month in months}
    monthly = frame.groupby("month")["expense_usd"].sum()
    return {month: round(float(monthly.get(month, 0.0)), 2) for month in months}


def _group_order(categories, expenses, report_group=None):
    if report_group:
        return [report_group]
    setup_groups = get_report_groups(categories)
    data_groups = sorted(expenses["report_group"].dropna().astype(str).unique().tolist()) if not expenses.empty else []
    return _ordered_unique(setup_groups + data_groups)


def _build_sections(transactions, categories, report_group=None):
    prepared_tx, expenses, months, month_labels = _prepare_report_data(transactions, categories, report_group)
    groups = _group_order(categories, expenses, report_group)
    all_total = float(expenses["expense_usd"].sum()) if not expenses.empty else 0.0
    active_months = int((expenses.groupby("month")["expense_usd"].sum() > 0).sum()) if not expenses.empty else 0
    average_denominator = max(active_months, 1)

    sections = []
    summary_rows = []

    for group in groups:
        group_expenses = expenses[expenses["report_group"] == group].copy()
        if group_expenses.empty:
            continue

        group_total = float(group_expenses["expense_usd"].sum())
        category_totals = (
            group_expenses.groupby("category")["expense_usd"]
            .sum()
            .sort_values(ascending=False)
        )
        rows = []
        for category, category_total in category_totals.items():
            category_expenses = group_expenses[group_expenses["category"] == category].copy()
            subcategory_totals = (
                category_expenses.groupby("subcategory", dropna=False)["expense_usd"]
                .sum()
                .sort_values(ascending=False)
            )
            if subcategory_totals.empty:
                subcategory_totals = pd.Series({"": category_total})

            first = True
            for subcategory, _ in subcategory_totals.items():
                sub_expenses = category_expenses[category_expenses["subcategory"] == subcategory]
                rows.append({
                    "report_group": group,
                    "category": category if first else "",
                    "category_total": round(float(category_total), 2) if first else None,
                    "category_percent": float(category_total / all_total) if first and all_total else None,
                    "category_average": round(float(category_total) / average_denominator, 2) if first else None,
                    "subcategory": subcategory,
                    "months": _month_totals(sub_expenses, months),
                })
                first = False

        sections.append({
            "group": group,
            "rows": rows,
            "total": round(group_total, 2),
            "percent": float(group_total / all_total) if all_total else 0.0,
            "average": round(group_total / average_denominator, 2),
            "months": _month_totals(group_expenses, months),
        })
        summary_rows.append({
            "report_group": group,
            "total": round(group_total, 2),
            "percent": float(group_total / all_total) if all_total else 0.0,
            "average": round(group_total / average_denominator, 2),
        })

    return prepared_tx, sections, summary_rows, months, month_labels


def _build_report_frame(transactions, categories, report_group=None):
    prepared_tx, sections, _, months, month_labels = _build_sections(transactions, categories, report_group)
    rows = []
    for section in sections:
        for row in section["rows"]:
            if not row["category"]:
                continue
            out = {
                "Report group": section["group"],
                "Category": row["category"],
                "Total expenses for all months ever": row["category_total"] or 0.0,
                "% of category from total": row["category_percent"] or 0.0,
                "Average monthly": row["category_average"] or 0.0,
                "Subcategory": row["subcategory"],
            }
            for month in months:
                out[month_labels[month]] = row["months"].get(month, 0.0)
            rows.append(out)
        total_row = {
            "Report group": section["group"],
            "Category": f"total of {section['group']} ONLY",
            "Total expenses for all months ever": section["total"],
            "% of category from total": section["percent"],
            "Average monthly": section["average"],
            "Subcategory": "",
        }
        for month in months:
            total_row[month_labels[month]] = section["months"].get(month, 0.0)
        rows.append(total_row)
    return pd.DataFrame(rows), prepared_tx


def _style_report_sheet(ws, max_month_col, header_rows, section_rows, total_rows, summary_marker_rows):
    header_fill = PatternFill("solid", fgColor="E8EEF6")
    section_fill = PatternFill("solid", fgColor="DDEDE9")
    total_fill = PatternFill("solid", fgColor="FFF4CC")
    summary_fill = PatternFill("solid", fgColor="EAF2FF")
    thin = Side(style="thin", color="D7DEE8")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    widths = {
        "A": 24,
        "B": 34,
        "C": 24,
        "D": 18,
        "E": 18,
        "F": 4,
        "G": 34,
    }
    for column, width in widths.items():
        ws.column_dimensions[column].width = width
    for col in range(8, max_month_col + 1):
        ws.column_dimensions[chr(64 + col)].width = 13

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
            if cell.row in header_rows:
                cell.fill = header_fill
                cell.font = Font(bold=True)
            if cell.row in section_rows:
                cell.fill = section_fill
                cell.font = Font(bold=True)
            if cell.row in total_rows:
                cell.fill = total_fill
                cell.font = Font(bold=True)
            if cell.row in summary_marker_rows:
                cell.fill = summary_fill
                cell.font = Font(bold=True)
            if cell.column in {3, 5} or cell.column >= 8:
                cell.number_format = '#,##0.00'
            if cell.column == 4:
                cell.number_format = '0.00%'

    ws.freeze_panes = "A4"


def _write_report_sheet(ws, title, sections, summary_rows, months, month_labels):
    header_rows = set()
    section_rows = set()
    total_rows = set()
    summary_marker_rows = set()

    month_headers = [month_labels[month] for month in months]
    headers = [
        "Report group",
        "Expense category",
        "Total expenses for all months ever",
        "% of category from total",
        "Average monthly",
        "",
        "Subcategory",
        *month_headers,
    ]
    max_col = len(headers)

    row_num = 1
    ws.cell(row_num, 1, title)
    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=max_col)
    ws.cell(row_num, 1).font = Font(bold=True, size=14)
    ws.cell(row_num, 1).alignment = Alignment(vertical="top", wrap_text=True)
    section_rows.add(row_num)
    row_num += 2

    if not sections:
        ws.cell(row_num, 1, "No reviewed expenses match the selected report filters.")
        return

    for section in sections:
        ws.cell(row_num, 1, section["group"])
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=max_col)
        section_rows.add(row_num)
        row_num += 1

        for col_num, header in enumerate(headers, 1):
            ws.cell(row_num, col_num, header)
        header_rows.add(row_num)
        row_num += 1

        for item in section["rows"]:
            values = [
                item["report_group"],
                item["category"],
                item["category_total"],
                item["category_percent"],
                item["category_average"],
                "",
                item["subcategory"],
                *[item["months"].get(month, 0.0) for month in months],
            ]
            for col_num, value in enumerate(values, 1):
                ws.cell(row_num, col_num, value)
            row_num += 1

        total_values = [
            section["group"],
            f"total of {section['group']} ONLY",
            section["total"],
            section["percent"],
            section["average"],
            "",
            "total",
            *[section["months"].get(month, 0.0) for month in months],
        ]
        for col_num, value in enumerate(total_values, 1):
            ws.cell(row_num, col_num, value)
        total_rows.add(row_num)
        row_num += 3

    ws.cell(row_num, 1, "Summary - total of all categories")
    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=5)
    summary_marker_rows.add(row_num)
    row_num += 1
    summary_headers = ["Report group", "Total", "% of category from total", "Average monthly"]
    for col_num, header in enumerate(summary_headers, 1):
        ws.cell(row_num, col_num, header)
    header_rows.add(row_num)
    row_num += 1
    for item in summary_rows:
        values = [item["report_group"], item["total"], item["percent"], item["average"]]
        for col_num, value in enumerate(values, 1):
            ws.cell(row_num, col_num, value)
        row_num += 1
    if summary_rows:
        grand_total = round(sum(float(item.get("total") or 0) for item in summary_rows), 2)
        grand_average = round(sum(float(item.get("average") or 0) for item in summary_rows), 2)
        total_values = ["TOTAL", grand_total, 1.0, grand_average]
        for col_num, value in enumerate(total_values, 1):
            ws.cell(row_num, col_num, value)
        total_rows.add(row_num)

    _style_report_sheet(ws, max_col, header_rows, section_rows, total_rows, summary_marker_rows)


def _prepare_verification_data(transactions, categories, report_group=None):
    tx = transactions.copy()
    original_count = len(tx)
    for column in [
        "id",
        "txn_date",
        "account_name",
        "bank",
        "account_number",
        "currency",
        "amount",
        "amount_usd",
        "category",
        "subcategory",
        "original_description",
    ]:
        if column not in tx.columns:
            tx[column] = ""

    tx["database_row"] = range(1, original_count + 1)
    tx["parsed_date"] = pd.to_datetime(tx["txn_date"], errors="coerce")
    tx["statement_amount_numeric"] = pd.to_numeric(tx["amount"], errors="coerce")
    tx["amount_usd_numeric"] = pd.to_numeric(tx["amount_usd"], errors="coerce")
    tx["currency_normalized"] = tx["currency"].fillna("").astype(str).str.upper().str.strip()
    tx["zero_usd_needs_attention"] = (
        tx["amount_usd_numeric"].fillna(0).abs().le(0.005)
        & tx["statement_amount_numeric"].fillna(0).abs().gt(0.005)
    )
    tx["report_amount"] = tx["amount_usd_numeric"]
    tx.loc[tx["zero_usd_needs_attention"] & ~tx["currency_normalized"].eq("USD"), "report_amount"] = pd.NA
    usd_fallback = (
        (tx["report_amount"].isna() | tx["zero_usd_needs_attention"])
        & tx["currency_normalized"].eq("USD")
        & tx["statement_amount_numeric"].notna()
    )
    tx.loc[usd_fallback, "report_amount"] = tx.loc[usd_fallback, "statement_amount_numeric"]
    tx["amount_source"] = "missing"
    tx.loc[
        tx["statement_amount_numeric"].notna()
        & tx["amount_usd_numeric"].isna()
        & ~tx["currency_normalized"].eq("USD"),
        "amount_source",
    ] = "missing USD equivalent"
    tx.loc[
        tx["zero_usd_needs_attention"]
        & ~tx["currency_normalized"].eq("USD"),
        "amount_source",
    ] = "zero USD equivalent"
    tx.loc[
        tx["amount_usd_numeric"].notna()
        & ~tx["zero_usd_needs_attention"],
        "amount_source",
    ] = "USD equivalent"
    tx.loc[usd_fallback, "amount_source"] = "statement amount (USD)"
    tx["report_amount"] = tx["report_amount"].fillna(0)

    group_map = _category_group_map(categories)
    tx["category"] = tx["category"].fillna("").astype(str).str.strip()
    tx["subcategory"] = tx["subcategory"].fillna("").astype(str).str.strip()
    tx["report_group"] = tx["category"].map(lambda value: group_map.get(str(value).casefold(), ""))
    tx["report_group"] = tx["report_group"].replace("", UNASSIGNED_GROUP)

    tx["in_report_group_scope"] = True if not report_group else tx["report_group"].eq(report_group)
    tx["valid_for_report"] = tx["parsed_date"].notna() & (
        (tx["amount_usd_numeric"].notna() & ~tx["zero_usd_needs_attention"])
        | (tx["currency_normalized"].eq("USD") & tx["statement_amount_numeric"].notna())
    )
    tx["is_own_funds"] = tx["category"].map(_is_own_funds)
    tx["income_like_category"] = tx.apply(
        lambda row: _looks_like_income_or_deposit(
            row.get("category", ""),
            row.get("subcategory", ""),
            row.get("report_group", ""),
        ),
        axis=1,
    )
    tx["is_income_or_deposit"] = (
        (tx["report_amount"] > 0)
        & ~tx["is_own_funds"]
        & (
            tx["income_like_category"]
            | tx["report_group"].eq(UNASSIGNED_GROUP)
        )
    )
    tx["is_expense"] = (
        ~tx["is_own_funds"]
        & (
            (tx["report_amount"] < 0)
            | (
                (tx["report_amount"] > 0)
                & ~tx["is_income_or_deposit"]
                & tx["report_group"].ne(UNASSIGNED_GROUP)
            )
        )
    )
    tx["expense_report_included"] = (
        tx["in_report_group_scope"]
        & tx["valid_for_report"]
        & tx["is_expense"]
    )
    tx["income_deposit_included"] = (
        tx["in_report_group_scope"]
        & tx["valid_for_report"]
        & tx["is_income_or_deposit"]
    )
    tx["own_funds_included"] = (
        tx["in_report_group_scope"]
        & tx["valid_for_report"]
        & tx["is_own_funds"]
    )
    tx["represented_in_workbook"] = tx["in_report_group_scope"] & tx["valid_for_report"]

    def status(row):
        if not row["in_report_group_scope"]:
            return "Outside selected reporting group"
        if pd.isna(row["parsed_date"]):
            return "Needs attention - missing or invalid date"
        if not row["valid_for_report"]:
            return "Needs attention - missing or zero USD equivalent/rate"
        if row["expense_report_included"]:
            return "Included in expense report"
        if row["income_deposit_included"]:
            return "Shown in Income deposits sheet"
        if row["own_funds_included"]:
            return "Shown in Own funds sheet - excluded from expense totals"
        return "Shown in verification only - zero amount"

    def section(row):
        if row["expense_report_included"]:
            return row["report_group"]
        if row["income_deposit_included"]:
            return "Income deposits"
        if row["own_funds_included"]:
            return "Own funds"
        return "Report verification"

    tx["report_status"] = tx.apply(status, axis=1)
    tx["report_section"] = tx.apply(section, axis=1)
    return tx


def build_report_verification(transactions, categories, report_group=None):
    tx = _prepare_verification_data(transactions, categories, report_group)
    in_scope = tx["in_report_group_scope"] if not tx.empty else pd.Series(dtype=bool)
    valid_in_scope = in_scope & tx["valid_for_report"] if not tx.empty else pd.Series(dtype=bool)

    expense_mask = tx["expense_report_included"] if not tx.empty else pd.Series(dtype=bool)
    deposit_mask = tx["income_deposit_included"] if not tx.empty else pd.Series(dtype=bool)
    own_funds_mask = tx["own_funds_included"] if not tx.empty else pd.Series(dtype=bool)
    represented_mask = tx["represented_in_workbook"] if not tx.empty else pd.Series(dtype=bool)
    attention_mask = in_scope & ~tx["valid_for_report"] if not tx.empty else pd.Series(dtype=bool)

    summary = {
        "database_rows": int(in_scope.sum()) if not tx.empty else 0,
        "represented_rows": int(represented_mask.sum()) if not tx.empty else 0,
        "expense_rows_in_report": int(expense_mask.sum()) if not tx.empty else 0,
        "deposit_rows": int(deposit_mask.sum()) if not tx.empty else 0,
        "own_funds_rows": int(own_funds_mask.sum()) if not tx.empty else 0,
        "rows_needing_attention": int(attention_mask.sum()) if not tx.empty else 0,
        "total_expenses": round(float(tx.loc[expense_mask, "report_amount"].abs().sum()), 2) if not tx.empty else 0.0,
        "total_deposits": round(float(tx.loc[deposit_mask, "report_amount"].sum()), 2) if not tx.empty else 0.0,
        "total_own_funds": round(float(tx.loc[own_funds_mask, "report_amount"].sum()), 2) if not tx.empty else 0.0,
        "net_movement": round(float(tx.loc[valid_in_scope, "report_amount"].sum()), 2) if not tx.empty else 0.0,
    }

    detail_columns = [
        ("id", "Database ID"),
        ("database_row", "Database row"),
        ("txn_date", "Date"),
        ("account_name", "Account"),
        ("bank", "Bank"),
        ("account_number", "Account number"),
        ("currency", "Currency"),
        ("amount", "Statement amount"),
        ("amount_usd", "USD equivalent"),
        ("report_amount", "USD amount used in report"),
        ("amount_source", "Amount source"),
        ("category", "Category"),
        ("subcategory", "Subcategory"),
        ("report_group", "Reporting group"),
        ("report_status", "Report status"),
        ("report_section", "Report workbook section"),
        ("expense_report_included", "Included in expense totals"),
        ("represented_in_workbook", "Represented in workbook"),
        ("original_description", "Full statement description"),
    ]
    detail = tx[[column for column, _ in detail_columns]].rename(
        columns={column: label for column, label in detail_columns}
    )
    return summary, detail


def _build_income_deposits_frame(verification_detail):
    if verification_detail.empty:
        return pd.DataFrame(columns=[
            "Date",
            "Account",
            "Bank",
            "Category",
            "Subcategory",
            "USD equivalent",
            "Full statement description",
        ])

    income = verification_detail[
        verification_detail["Report status"].eq("Shown in Income deposits sheet")
    ].copy()
    if income.empty:
        return pd.DataFrame([{
            "Date": "",
            "Account": "",
            "Bank": "",
            "Category": "No income/deposit rows in the selected report data",
            "Subcategory": "",
            "USD equivalent": 0.0,
            "Full statement description": "",
        }])

    income = income[[
        "Date",
        "Account",
        "Bank",
        "Category",
        "Subcategory",
        "USD amount used in report",
        "Full statement description",
    ]].rename(columns={"USD amount used in report": "USD equivalent"})
    total = round(float(pd.to_numeric(income["USD equivalent"], errors="coerce").fillna(0).sum()), 2)
    total_row = {
        "Date": "",
        "Account": "",
        "Bank": "",
        "Category": "TOTAL INCOME / DEPOSITS",
        "Subcategory": "",
        "USD equivalent": total,
        "Full statement description": "",
    }
    return pd.concat([income, pd.DataFrame([total_row])], ignore_index=True)


def _build_own_funds_frame(verification_detail):
    if verification_detail.empty:
        return pd.DataFrame(columns=[
            "Date",
            "Account",
            "Bank",
            "Category",
            "Subcategory",
            "USD equivalent",
            "Full statement description",
        ])

    own_funds = verification_detail[
        verification_detail["Report status"].eq("Shown in Own funds sheet - excluded from expense totals")
    ].copy()
    if own_funds.empty:
        return pd.DataFrame([{
            "Date": "",
            "Account": "",
            "Bank": "",
            "Category": "No own funds rows in the selected report data",
            "Subcategory": "",
            "USD equivalent": 0.0,
            "Full statement description": "",
        }])

    own_funds = own_funds[[
        "Date",
        "Account",
        "Bank",
        "Category",
        "Subcategory",
        "USD amount used in report",
        "Full statement description",
    ]].rename(columns={"USD amount used in report": "USD equivalent"})
    total = round(float(pd.to_numeric(own_funds["USD equivalent"], errors="coerce").fillna(0).sum()), 2)
    total_row = {
        "Date": "",
        "Account": "",
        "Bank": "",
        "Category": "TOTAL OWN FUNDS",
        "Subcategory": "",
        "USD equivalent": total,
        "Full statement description": "",
    }
    return pd.concat([own_funds, pd.DataFrame([total_row])], ignore_index=True)


def _verification_summary_frame(summary):
    return pd.DataFrame([
        {"Metric": "Database rows checked", "Value": summary["database_rows"]},
        {"Metric": "Rows represented in workbook", "Value": summary["represented_rows"]},
        {"Metric": "Expense rows included in expense totals", "Value": summary["expense_rows_in_report"]},
        {"Metric": "Income/deposit rows shown separately", "Value": summary["deposit_rows"]},
        {"Metric": "Own funds rows shown separately", "Value": summary["own_funds_rows"]},
        {"Metric": "Rows needing attention", "Value": summary["rows_needing_attention"]},
        {"Metric": "Total expenses in report", "Value": summary["total_expenses"]},
        {"Metric": "Total income/deposits", "Value": summary["total_deposits"]},
        {"Metric": "Total own funds", "Value": summary["total_own_funds"]},
        {"Metric": "Net movement", "Value": summary["net_movement"]},
    ])


def _style_table_sheet(ws, wrap_columns=None):
    wrap_columns = wrap_columns or set()
    header_fill = PatternFill("solid", fgColor="E8EEF6")
    thin = Side(style="thin", color="D7DEE8")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = border

    for column_cells in ws.columns:
        column_letter = get_column_letter(column_cells[0].column)
        header = str(column_cells[0].value or "")
        max_len = max(len(str(cell.value or "")) for cell in column_cells[:80])
        ws.column_dimensions[column_letter].width = min(max(max_len + 2, 12), 44)
        if header in wrap_columns or "description" in header.casefold():
            ws.column_dimensions[column_letter].width = 58
        for cell in column_cells:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
            if isinstance(cell.value, float):
                cell.number_format = '#,##0.00'
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def build_sample_expenses_report(transactions, categories, report_group=None):
    output = BytesIO()
    prepared_tx, sections, summary_rows, months, month_labels = _build_sections(transactions, categories, report_group)
    verification_summary, verification_detail = build_report_verification(transactions, categories, report_group)
    income_deposits = _build_income_deposits_frame(verification_detail)
    own_funds = _build_own_funds_frame(verification_detail)
    title = "Sample expenses report" if not report_group else f"Sample expenses report - {report_group}"

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        prepared_tx.to_excel(writer, index=False, sheet_name="Reviewed transactions")
        income_deposits.to_excel(writer, index=False, sheet_name="Income deposits")
        own_funds.to_excel(writer, index=False, sheet_name="Own funds")
        _verification_summary_frame(verification_summary).to_excel(writer, index=False, sheet_name="Report check")
        verification_detail.to_excel(writer, index=False, sheet_name="Report verification")
        ws = writer.book.create_sheet("Sample expenses report", 0)
        _write_report_sheet(ws, title, sections, summary_rows, months, month_labels)
        for sheet_name in ["Reviewed transactions", "Income deposits", "Own funds", "Report check", "Report verification"]:
            if sheet_name in writer.book.sheetnames:
                _style_table_sheet(
                    writer.book[sheet_name],
                    wrap_columns={"Full statement description", "Report status"},
                )
    return output.getvalue()


def build_excel_report(context):
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        context.get("transactions", pd.DataFrame()).to_excel(writer, index=False, sheet_name="Transactions")
    return output.getvalue()


def build_csv_report(context):
    return context.get("transactions", pd.DataFrame()).to_csv(index=False).encode("utf-8")


def _pdf_safe(text):
    return str(text).replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)")


def _format_money(value):
    if value in (None, ""):
        return ""
    try:
        return f"{float(value):,.2f}"
    except Exception:
        return str(value)


def _format_percent(value):
    if value in (None, ""):
        return ""
    try:
        return f"{float(value) * 100:.2f}%"
    except Exception:
        return str(value)


def _pdf_lines_for_sections(sections, summary_rows, title, months, month_labels):
    grand_total = round(sum(float(item.get("total") or 0) for item in summary_rows), 2)
    grand_average = round(sum(float(item.get("average") or 0) for item in summary_rows), 2)
    lines = [
        title,
        "Reports exclude Own funds and use USD equivalent where available.",
        f"Total expenses in this report: {_format_money(grand_total)}",
        "",
    ]
    if not sections:
        return lines + ["No reviewed expenses match the selected report filters."]

    for section in sections:
        lines.append(section["group"])
        lines.append("Category | Total | % of total | Average monthly | Subcategory")
        lines.append("-" * 120)
        for item in section["rows"]:
            if item["category"]:
                line = (
                    f"{item['category']} | {_format_money(item['category_total'])} | "
                    f"{_format_percent(item['category_percent'])} | {_format_money(item['category_average'])} | "
                    f"{item['subcategory']}"
                )
            else:
                line = f" |  |  |  | {item['subcategory']}"
            lines.extend(textwrap.wrap(line, width=150) or [""])
            month_values = [
                f"{month_labels[month]}: {_format_money(item['months'].get(month, 0.0))}"
                for month in months
                if item["months"].get(month, 0.0)
            ]
            if month_values:
                lines.extend(textwrap.wrap("Months: " + ", ".join(month_values), width=150))
        total_line = (
            f"total of {section['group']} ONLY | {_format_money(section['total'])} | "
            f"{_format_percent(section['percent'])} | {_format_money(section['average'])}"
        )
        lines.extend(["", total_line, ""])

    lines.append("Summary - total of all categories")
    for item in summary_rows:
        lines.append(
            f"{item['report_group']} | {_format_money(item['total'])} | "
            f"{_format_percent(item['percent'])} | {_format_money(item['average'])}"
        )
    if summary_rows:
        lines.append(f"TOTAL | {_format_money(grand_total)} | 100.00% | {_format_money(grand_average)}")
    return lines


def _minimal_pdf(lines):
    width, height = 842, 595
    lines_per_page = 44
    pages = [lines[i:i + lines_per_page] for i in range(0, len(lines), lines_per_page)] or [[]]
    objects = []

    def add(obj):
        objects.append(obj)
        return len(objects)

    catalog_id = add("<< /Type /Catalog /Pages 2 0 R >>")
    pages_id = add("")
    page_ids = []
    font_id = add("<< /Type /Font /Subtype /Type1 /BaseFont /Courier >>")

    for page_lines in pages:
        content = ["BT", "/F1 7 Tf", "36 560 Td", "10 TL"]
        for line in page_lines:
            content.append(f"({_pdf_safe(line[:190])}) Tj")
            content.append("T*")
        content.append("ET")
        stream = "\n".join(content)
        content_id = add(f"<< /Length {len(stream.encode('latin-1', errors='ignore'))} >>\nstream\n{stream}\nendstream")
        page_id = add(
            f"<< /Type /Page /Parent {pages_id} 0 R /MediaBox [0 0 {width} {height}] "
            f"/Resources << /Font << /F1 {font_id} 0 R >> >> /Contents {content_id} 0 R >>"
        )
        page_ids.append(page_id)

    objects[pages_id - 1] = (
        f"<< /Type /Pages /Kids [{' '.join(f'{page_id} 0 R' for page_id in page_ids)}] "
        f"/Count {len(page_ids)} >>"
    )

    output = BytesIO()
    output.write(b"%PDF-1.4\n")
    offsets = [0]
    for idx, obj in enumerate(objects, start=1):
        offsets.append(output.tell())
        output.write(f"{idx} 0 obj\n{obj}\nendobj\n".encode("latin-1", errors="ignore"))
    xref = output.tell()
    output.write(f"xref\n0 {len(objects) + 1}\n".encode("ascii"))
    output.write(b"0000000000 65535 f \n")
    for offset in offsets[1:]:
        output.write(f"{offset:010d} 00000 n \n".encode("ascii"))
    output.write(
        f"trailer\n<< /Size {len(objects) + 1} /Root {catalog_id} 0 R >>\n"
        f"startxref\n{xref}\n%%EOF".encode("ascii")
    )
    return output.getvalue()


def safe_filename(value):
    return re.sub(r"[^A-Za-z0-9._-]+", "_", str(value)).strip("_") or "report"


def build_pdf_report(transactions, categories, report_group=None):
    title = "Sample expenses report" if not report_group else f"Sample expenses report - {report_group}"
    _, sections, summary_rows, months, month_labels = _build_sections(transactions, categories, report_group)
    verification_summary, _ = build_report_verification(transactions, categories, report_group)
    lines = _pdf_lines_for_sections(sections, summary_rows, title, months, month_labels)
    lines.extend([
        "",
        "Report verification",
        f"Database rows checked: {verification_summary['database_rows']}",
        f"Rows represented in workbook: {verification_summary['represented_rows']}",
        f"Income/deposit rows shown separately: {verification_summary['deposit_rows']}",
        f"Own funds rows shown separately: {verification_summary['own_funds_rows']}",
        f"Rows needing attention: {verification_summary['rows_needing_attention']}",
    ])
    return _minimal_pdf(lines)
